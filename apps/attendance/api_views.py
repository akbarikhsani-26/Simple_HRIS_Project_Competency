import csv
import datetime

import holidays
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import filters, mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.employees.models import Employee

from .models import Attendance
from .serializers import AttendanceSerializer


class AttendanceViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ["karyawan__nama", "karyawan__jabatan"]

    def get_queryset(self):
        user = self.request.user
        if user.role == "HR_ADMIN":
            return Attendance.objects.all().order_by("-tanggal", "-jam_masuk")
        return Attendance.objects.filter(karyawan__user=user).order_by(
            "-tanggal", "-jam_masuk"
        )

    @action(detail=False, methods=["post"], url_path="check-in")
    def check_in(self, request):
        user = request.user

        # 1. Holiday & Weekend Validation
        today = timezone.localdate()
        id_holidays = holidays.ID()

        if today in id_holidays:
            holiday_name = id_holidays.get(today)
            return Response(
                {
                    "error": f"Hari ini adalah hari libur nasional ({holiday_name}). Anda tidak dapat melakukan absensi."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if today.weekday() >= 5:  # 5 = Saturday, 6 = Sunday
            return Response(
                {
                    "error": "Hari ini adalah akhir pekan. Anda tidak dapat melakukan absensi."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 2. Time Cutoff Validation (08:00)
        current_time = timezone.localtime().time()
        cutoff_time = datetime.time(8, 0, 0)

        status_kehadiran = (
            "TERLAMBAT" if current_time > cutoff_time else "TEPAT_WAKTU"
        )

        try:
            karyawan = user.employee_profile
        except Employee.DoesNotExist:
            if user.role == "HR_ADMIN":
                karyawan = Employee.objects.create(
                    user=user,
                    nama=user.email.split("@")[0].capitalize(),
                    email=user.email,
                    jabatan="HR Admin",
                )
            else:
                return Response(
                    {"error": "Profil karyawan tidak ditemukan."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if Attendance.objects.filter(
            karyawan=karyawan, tanggal=today
        ).exists():
            return Response(
                {"error": "Anda sudah melakukan check-in hari ini."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Create attendance record
        attendance = Attendance.objects.create(
            karyawan=karyawan, status=status_kehadiran
        )
        serializer = self.get_serializer(attendance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="check-out")
    def check_out(self, request):
        user = request.user
        try:
            karyawan = user.employee_profile
        except Employee.DoesNotExist:
            return Response(
                {"error": "Profil karyawan tidak ditemukan."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today = timezone.localdate()

        try:
            attendance = Attendance.objects.get(
                karyawan=karyawan, tanggal=today
            )
        except Attendance.DoesNotExist:
            return Response(
                {"error": "Anda belum melakukan check-in hari ini."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if attendance.jam_keluar is not None:
            return Response(
                {"error": "Anda sudah melakukan check-out hari ini."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        attendance.jam_keluar = timezone.localtime().time()
        attendance.save()

        serializer = self.get_serializer(attendance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="export-csv")
    def export_csv(self, request):
        if request.user.role != "HR_ADMIN":
            return Response(
                {"error": "Hanya HR_ADMIN yang dapat mengunduh laporan."},
                status=status.HTTP_403_FORBIDDEN,
            )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            'attachment; filename="laporan_absensi.csv"'
        )

        writer = csv.writer(response)
        writer.writerow(
            [
                "Nama Karyawan",
                "Email",
                "Jabatan",
                "Tanggal",
                "Jam Masuk",
                "Jam Keluar",
                "Status",
            ]
        )

        today = timezone.localdate()
        search_query = request.GET.get("search", "")
        date_query = request.GET.get("date", "")
        status_query = request.GET.get("status", "")

        if not date_query:
            target_date = today
        else:
            try:
                target_date = datetime.datetime.strptime(
                    date_query, "%Y-%m-%d"
                ).date()
            except ValueError:
                target_date = today

        employees = Employee.objects.filter(status_aktif=True)
        if search_query:
            employees = employees.filter(nama__icontains=search_query)

        attendances = {
            att.karyawan_id: att
            for att in Attendance.objects.filter(tanggal=target_date)
        }

        if not employees.exists():
            writer.writerow(
                ["Tidak ada data karyawan", "", "", "", "", "", ""]
            )
            return response

        for emp in employees:
            att = attendances.get(emp.id)
            status_display = att.get_status_display() if att else "Tidak Masuk"

            if status_query:
                if status_query == "ALPA" and status_display != "Tidak Masuk":
                    continue
                if (
                    status_query == "TEPAT_WAKTU"
                    and status_display != "Tepat Waktu"
                ):
                    continue
                if (
                    status_query == "TERLAMBAT"
                    and status_display != "Terlambat"
                ):
                    continue

            if att:
                writer.writerow(
                    [
                        emp.nama,
                        emp.email,
                        emp.jabatan,
                        att.tanggal.strftime("%Y-%m-%d"),
                        (
                            att.jam_masuk.strftime("%H:%M:%S")
                            if att.jam_masuk
                            else "-"
                        ),
                        (
                            att.jam_keluar.strftime("%H:%M:%S")
                            if att.jam_keluar
                            else "-"
                        ),
                        status_display,
                    ]
                )
            else:
                writer.writerow(
                    [
                        emp.nama,
                        emp.email,
                        emp.jabatan,
                        target_date.strftime("%Y-%m-%d"),
                        "-",
                        "-",
                        status_display,
                    ]
                )

        return response

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        if request.user.role != "HR_ADMIN":
            return Response(
                {"error": "Hanya HR_ADMIN yang dapat mengunduh laporan."},
                status=status.HTTP_403_FORBIDDEN,
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="laporan_absensi.xlsx"'
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Laporan Absensi"

        columns = [
            "Nama Karyawan",
            "Email",
            "Jabatan",
            "Tanggal",
            "Jam Masuk",
            "Jam Keluar",
            "Status",
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        today = timezone.localdate()
        search_query = request.GET.get("search", "")
        date_query = request.GET.get("date", "")
        status_query = request.GET.get("status", "")

        if not date_query:
            target_date = today
        else:
            try:
                target_date = datetime.datetime.strptime(
                    date_query, "%Y-%m-%d"
                ).date()
            except ValueError:
                target_date = today

        employees = Employee.objects.filter(status_aktif=True)
        if search_query:
            employees = employees.filter(nama__icontains=search_query)

        attendances = {
            att.karyawan_id: att
            for att in Attendance.objects.filter(tanggal=target_date)
        }

        if not employees.exists():
            worksheet.cell(row=2, column=1).value = "Tidak ada data karyawan"
        else:
            for emp in employees:
                att = attendances.get(emp.id)
                status_display = (
                    att.get_status_display() if att else "Tidak Masuk"
                )

                if status_query:
                    if (
                        status_query == "ALPA"
                        and status_display != "Tidak Masuk"
                    ):
                        continue
                    if (
                        status_query == "TEPAT_WAKTU"
                        and status_display != "Tepat Waktu"
                    ):
                        continue
                    if (
                        status_query == "TERLAMBAT"
                        and status_display != "Terlambat"
                    ):
                        continue

                row_num += 1
                if att:
                    worksheet.cell(row=row_num, column=1).value = emp.nama
                    worksheet.cell(row=row_num, column=2).value = emp.email
                    worksheet.cell(row=row_num, column=3).value = emp.jabatan
                    worksheet.cell(row=row_num, column=4).value = (
                        att.tanggal.strftime("%Y-%m-%d")
                    )
                    worksheet.cell(row=row_num, column=5).value = (
                        att.jam_masuk.strftime("%H:%M:%S")
                        if att.jam_masuk
                        else "-"
                    )
                    worksheet.cell(row=row_num, column=6).value = (
                        att.jam_keluar.strftime("%H:%M:%S")
                        if att.jam_keluar
                        else "-"
                    )
                    worksheet.cell(row=row_num, column=7).value = (
                        status_display
                    )
                else:
                    worksheet.cell(row=row_num, column=1).value = emp.nama
                    worksheet.cell(row=row_num, column=2).value = emp.email
                    worksheet.cell(row=row_num, column=3).value = emp.jabatan
                    worksheet.cell(row=row_num, column=4).value = (
                        target_date.strftime("%Y-%m-%d")
                    )
                    worksheet.cell(row=row_num, column=5).value = "-"
                    worksheet.cell(row=row_num, column=6).value = "-"
                    worksheet.cell(row=row_num, column=7).value = (
                        status_display
                    )

        workbook.save(response)
        return response
